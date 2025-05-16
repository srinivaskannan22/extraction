from PIL import Image
from io import BytesIO
import chardet
from fastapi.responses import JSONResponse
import pandas as pd
import os
from openpyxl import load_workbook
import docx2txt
import pdfplumber
from pdfplumber.utils import extract_text, get_bbox_overlap, obj_to_bbox
import time
import pytesseract
from google import genai
from google.genai.types import HttpOptions, Part
from google.genai import types
import boto3
import json
import openpyxl
from dotenv import load_dotenv
from fastapi import HTTPException
load_dotenv()

class Extraction:
    def __init__(self, path):
        self.path = path
    def image_extraction(self):
        image_stream = BytesIO(self.path)
        img = Image.open(image_stream)
        text = pytesseract.image_to_string(img)
        return {'status_code':200,"message":'extract the image using pytesseract',"data":text}
    def detect_encoding(self,file_byte):
        result= chardet.detect(file_byte)
        return result['encoding']
    def csv_extraction(self):
        encodings=[self.detect_encoding(self.path), 'utf-8', 'latin1', 'windows-1252']
        for enc in encodings:
            try:
                csv_=BytesIO(self.path)
                print (f'trying encoding : {enc}')
                data = pd.read_csv(csv_, encoding=enc)
                # print(data)
                data_cleaned = data.where(pd.notnull(data), None)
                print(data_cleaned.to_dict(orient='records'))
                '''
                format of the response body
                    {
                        "status": 2xx,
                        "message": ""
                        "data": []/{}
                    }
                '''
                return {"status_code":200,"message":f"CSV loaded successfully using encoding: {enc}","data":data_cleaned.to_dict(orient='records')} 
            except Exception as error:
                raise HTTPException(status_code=500,detail=f"error occured {error}")
    def  docx_extraction(self):
        path2=BytesIO(self.path)
        
        text = docx2txt.process(path2)
        return {"status_code":200,"message":"docs extract successfully","data":text} 
    
    def extract_text_with_fallback(self,page):
        """
        Extract text from a PDF page. If extraction fails, fall back to OCR.
        """
        try:
            text = page.extract_text()
            if text:
                return text
            else:
                img = page.to_image(resolution=300).original  
                text = pytesseract.image_to_string(img)
                return text
        except Exception as e:
            raise HTTPException(status_code=500,detail=f'error need due to{e}')

    def process_pdf(self):
        pdf_path=BytesIO(self.path)

        start_time = time.time()
        pdf = pdfplumber.open(pdf_path)
        all_text = []
        main_table_headers = None
        main_table_parts = []

        for page_num, page in enumerate(pdf.pages):
            filtered_page = page
            chars = filtered_page.chars
            tables = page.find_tables()
            for table in tables:
                table_bbox = table.bbox
                table_cropped = page.crop(table_bbox)
                table_data = table.extract()
                if not table_data:
                    continue

                if main_table_headers is None:
                    if len(table_data) > 1:
                        main_table_headers = table_data[0]

                        main_table_data = [[cell if cell is not None else '' for cell in row] for row in table_data[1:]]
                        main_table_parts.extend(main_table_data)
                else:
                    if len(table_data[0]) == len(main_table_headers):
                        for row in table_data:
                            row = [cell if cell is not None else '' for cell in row]
                            main_table_parts.append(row)
                filtered_page = filtered_page.filter(lambda obj: 
                    get_bbox_overlap(obj_to_bbox(obj), table_bbox) is None
                )
            page_text = self.extract_text_with_fallback(filtered_page)
            all_text.append(page_text)
        final_main_table = []
        current_row = None
        for row in main_table_parts:
            if row[0] != '' or len([cell for cell in row if cell != '']) > 1:
                if current_row is not None:
                    final_main_table.append(current_row)
                current_row = row
            else:
                if current_row is not None:
                    for col_index, cell in enumerate(row):
                        if cell != '':
                            if current_row[col_index]:
                                current_row[col_index] += ' ' + cell
                            else:
                                current_row[col_index] = cell
                else:
                    current_row = row
        if current_row is not None:
            final_main_table.append(current_row)

        if main_table_headers and final_main_table:
            df = pd.DataFrame(final_main_table, columns=main_table_headers)
            markdown_table = df.to_markdown(index=False)
            all_text.append(markdown_table)
        pdf.close()
        # return "\n".join(all_text)
        return {"status_code":200,"message":"docs extract successfully","data":"\n".join(all_text)} 
    def extract_merged_cells_from_excel(self,ws):
        data = []
        for row in ws.iter_rows():
            data.append([cell.value if cell.value is not None else "" for cell in row])
        
        df = pd.DataFrame(data)
        for col in df.columns:  
            prev_value = None  
            for idx in df.index:  
                current_value = df.at[idx, col] 
                if current_value == "":  
                    continue  
                else:
                    prev_value = current_value  
        return df

    def process_excel_file(self):
        path=BytesIO(self.path)
        
        wb = load_workbook(path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"Processing sheet: {ws}")
            df = self.extract_merged_cells_from_excel(ws)
            print(f"Data from sheet '{sheet_name}' extracted successfully!")
            print(df)
            # output_path = os.path.join(output_dir, f"{sheet_name}.csv")
            # print(f"Saving data from sheet '{sheet_name}' to {output_path}")
            # df.to_csv(output_path, index=False, header=False)
        # return df 
        return {"status_code":200,"message":" openpyxl extract successfully","data":df}   
    
    def gemini_extraction(self):
        try:
            os.environ['GOOGLE_APPLICATION_CREDENTIALS']=os.getenv('GOOGLE_APPLICATION_CREDENTIALS')
            client = genai.Client(vertexai=True,project="bootlabs-ai-397504",location='us-central1')
            content = [
                types.Part(text="this is bill so extract the relevant-content and given me in the form of json , "),
                types.Part.from_bytes(data=self.path, mime_type="image/jpeg")
            ]

            response = client.models.generate_content(
                model='gemini-2.0-flash-001',
                contents=content
                )
            final =response.text
            cleaned_response = final.replace("```json\n", "").replace("```", "").strip()
            json_response=json.loads(cleaned_response)
            return {"status_code":200,"message":"gemini extract successfully","data":json_response} 
        except Exception as error :
            print(f'error due to {error}')
            raise HTTPException(status_code=500,detail= f'error due {error}')
    def amazon_texract(self):
        try:
            client = boto3.client(
                        'textract',
                        aws_access_key_id=os.getenv('aws_access_key_id'),
                        aws_secret_access_key=os.getenv("aws_secret_access_key"),
                        region_name="us-east-1"
                    )
            response=client.analyze_document(Document={"Bytes":self.path},
                                    FeatureTypes=['FORMS'])
            extraction=[]
            for block in response['Blocks']:
                if block['BlockType']=="LINE":
                        extraction.append((block['Text']))
            return {"status_code":200,"message":"texeract extract successfully","data":extraction} 
        except Exception as err:
            raise HTTPException(status_code=500,detail=f'error due to {err}')

    def invocie_agent(self):
        try:
            summary_entities_values = []
            summary_fields = []
            expense_item = []

            client=boto3.client('textract')
            response=client.analyze_expense(Document={
            "Bytes":self.path,
            
            })

            for expense_doc in response["ExpenseDocuments"]:
                for field in expense_doc["SummaryFields"]:
                    kvs = {}
                    if "LabelDetection" in field:
                        if "ValueDetection" in field:
                            kvs[field["LabelDetection"]["Text"]] = field["ValueDetection"]["Text"]
                    else:
                        kvs[field["Type"]["Text"]] = field["ValueDetection"]["Text"]
                    summary_entities_values.append(kvs.copy())
                    kvs = None

                for line_item_group in expense_doc["LineItemGroups"]:
                        for line_items in line_item_group["LineItems"]:
                            for field in line_items["LineItemExpenseFields"]:
                                kvs = {}
                                if "LabelDetection" in field:
                                    if "ValueDetection" in field:
                                        kvs[field["LabelDetection"]["Text"]] = field["ValueDetection"]["Text"]
                                else:
                                    kvs[field["Type"]["Text"]] = field["ValueDetection"]["Text"]
                                expense_item.append(kvs.copy())
                                kvs = None
            return {"status_code":200,"message":"texeract invoice  extract successfully","data":summary_entities_values}
        except Exception as error:
            raise HTTPException(status_code=500, detail=f'error due to{error}')  
    def _extract_excel(self):
        exc=BytesIO(self.path)
        wb=openpyxl.load_workbook(exc)
        sheetnames=wb.sheetnames
        extraction=[]
        for sheet in sheetnames:
            ws = wb[sheet]
            extraction.append(list(ws.values))
        return {"status_code":200,"message":"excel data is extract using openpyxl","data":extraction}     






        
                        
       


